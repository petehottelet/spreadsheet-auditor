"""SARIF output contract and schema validation."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest

from spreadsheet_auditor.finding import Finding
from spreadsheet_auditor.report import build_payload
from spreadsheet_auditor.sarif import SARIF_VERSION, render_sarif


def _payload():
    workbook_meta = {
        "path": "demo.xlsx",
        "sha256": "abc",
        "sheets_analyzed": 1,
        "formulas_scanned": 1,
        "recalc_status": "completed",
    }
    coverage = {"macros_present": False, "macros_executed": False, "external_links_present": False, "unsupported_features": [], "limitations": []}
    findings = [
        Finding("BROKEN_REFERENCE", "High", "Defect", "DET", "Budget!B14", "broken", ["#REF!"], "fix", formula="=SUM(#REF!)"),
        Finding("IFERROR_MASK", "Medium", "Review", "HEUR", "Budget!C5", "mask", ["IFERROR"], "review", formula="=IFERROR(...)"),
    ]
    return build_payload("0.1.0", workbook_meta, coverage, findings)


def test_sarif_shape_is_2_1_0():
    sarif = render_sarif(_payload())
    assert sarif["version"] == SARIF_VERSION
    assert sarif["$schema"].endswith("sarif-schema-2.1.0.json")
    assert len(sarif["runs"]) == 1
    run = sarif["runs"][0]
    assert run["tool"]["driver"]["name"] == "spreadsheet-auditor"
    assert run["tool"]["driver"]["version"]
    assert run["results"]


def test_sarif_results_carry_severity_and_fingerprint():
    sarif = render_sarif(_payload())
    result = sarif["runs"][0]["results"][0]
    assert result["ruleId"] == "BROKEN_REFERENCE"
    assert result["level"] in {"error", "warning", "note"}
    assert "partialFingerprints" in result
    assert "spreadsheetAuditor/v1" in result["partialFingerprints"]
    assert result["properties"]["severity"] == "High"
    assert result["properties"]["sheet"] == "Budget"
    assert result["properties"]["cell"] == "B14"


def test_sarif_rules_are_unique():
    sarif = render_sarif(_payload())
    rules = sarif["runs"][0]["tool"]["driver"]["rules"]
    ids = [r["id"] for r in rules]
    assert len(ids) == len(set(ids))


def test_cli_format_sarif_writes_file(tmp_path):
    workbook = Path(__file__).resolve().parents[1] / "examples" / "demo_bad_budget.xlsx"
    if not workbook.exists():
        pytest.skip("demo workbook not generated yet")
    out = tmp_path / "report.sarif"
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "spreadsheet_auditor",
            str(workbook),
            "--out",
            str(out),
            "--format",
            "sarif",
            "--fail-on",
            "None",
        ],
        capture_output=True,
        text=True,
    )
    assert result.returncode in (0, 2), result.stderr
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["version"] == SARIF_VERSION
    assert data["runs"][0]["results"]


def test_sarif_validates_against_official_schema(tmp_path):
    """Validate generated SARIF against the bundled SARIF 2.1.0 schema.

    The schema is draft-07 and uses ``format: date-time``; we pass an explicit
    ``FormatChecker`` so that format validation is actually enforced rather than
    silently skipped.
    """
    jsonschema = pytest.importorskip("jsonschema")
    schema_path = Path(__file__).resolve().parents[1] / "schemas" / "sarif-2.1.0.schema.json"
    assert schema_path.exists(), "SARIF schema must be bundled under schemas/"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    sarif = render_sarif(_payload())
    jsonschema.validate(sarif, schema, format_checker=jsonschema.FormatChecker())


def test_write_report_sarif_does_not_fall_back_to_json(tmp_path, monkeypatch):
    """A SARIF rendering failure must propagate, never write generic JSON."""
    from spreadsheet_auditor import audit, sarif

    def _boom(*_args, **_kwargs):
        raise RuntimeError("sarif boom")

    monkeypatch.setattr(sarif, "write_sarif", _boom)
    out = tmp_path / "report.sarif"
    with pytest.raises(RuntimeError, match="sarif boom"):
        audit._write_report(_payload(), str(out), "sarif")
    assert not out.exists()


def test_cli_returns_5_when_sarif_write_fails(tmp_path, monkeypatch):
    """The CLI surfaces a SARIF write failure as exit code 5, not a fake success."""
    from openpyxl import Workbook

    from spreadsheet_auditor import audit, sarif

    workbook = tmp_path / "tiny.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=SUM(#REF!)"
    wb.save(workbook)

    def _boom(*_args, **_kwargs):
        raise RuntimeError("sarif boom")

    monkeypatch.setattr(sarif, "write_sarif", _boom)
    out = tmp_path / "report.sarif"
    code = audit.main(
        [str(workbook), "--out", str(out), "--format", "sarif", "--fail-on", "None"]
    )
    assert code == 5
    assert not out.exists()
