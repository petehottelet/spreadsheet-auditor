"""Unit tests for individual detectors and parsing helpers."""

from openpyxl import Workbook

from config_loader import check_setting
from formula_parser import ParsedReference, extract_numeric_literals
from range_checks import detect_range_length_mismatch
from reconcile import detect_cross_foot_failures
from reference_resolver import reference_in_bounds


def _wb(values: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for coord, value in values.items():
        ws[coord] = value
    return wb


def _formula_cell(row: int, col: int, formula: str) -> dict:
    from openpyxl.utils.cell import get_column_letter

    coord = f"{get_column_letter(col)}{row}"
    return {
        "sheet": "S",
        "row": row,
        "col": col,
        "coord": coord,
        "location": f"S!{coord}",
        "formula": formula,
    }


def test_range_length_mismatch_flags_short_peer():
    cells = [
        _formula_cell(1, 6, "=SUM(A1:C1)"),
        _formula_cell(2, 6, "=SUM(A2:C2)"),
        _formula_cell(3, 6, "=SUM(A3:B3)"),
    ]
    findings = detect_range_length_mismatch(cells)
    assert [f.location for f in findings] == ["S!F3"]


def test_range_length_mismatch_quiet_when_consistent():
    cells = [
        _formula_cell(1, 6, "=SUM(A1:C1)"),
        _formula_cell(2, 6, "=SUM(A2:C2)"),
        _formula_cell(3, 6, "=SUM(A3:C3)"),
    ]
    assert detect_range_length_mismatch(cells) == []


def test_cross_foot_detects_mismatch():
    formula_wb = _wb(
        {
            "E2": "=SUM(B2:D2)",
            "E3": "=SUM(B3:D3)",
            "B4": "=SUM(B2:B2)",
            "C4": "=SUM(C2:C3)",
            "D4": "=SUM(D2:D3)",
            "E4": "=SUM(E2:E3)",
        }
    )
    value_wb = _wb({"E2": 60, "E3": 45, "B4": 10, "C4": 35, "D4": 55, "E4": 105})
    findings = detect_cross_foot_failures(formula_wb, value_wb)
    assert any(f.rule_id == "CROSS_FOOT_FAILURE" and f.location == "S!E4" for f in findings)


def test_cross_foot_quiet_when_balanced():
    formula_wb = _wb(
        {
            "E2": "=SUM(B2:D2)",
            "E3": "=SUM(B3:D3)",
            "B4": "=SUM(B2:B3)",
            "C4": "=SUM(C2:C3)",
            "D4": "=SUM(D2:D3)",
            "E4": "=SUM(E2:E3)",
        }
    )
    value_wb = _wb({"E2": 60, "E3": 45, "B4": 15, "C4": 35, "D4": 55, "E4": 105})
    assert detect_cross_foot_failures(formula_wb, value_wb) == []


def test_cross_foot_skips_without_cached_values():
    formula_wb = _wb(
        {
            "E2": "=SUM(B2:D2)",
            "E3": "=SUM(B3:D3)",
            "B4": "=SUM(B2:B2)",
            "C4": "=SUM(C2:C3)",
            "D4": "=SUM(D2:D3)",
            "E4": "=SUM(E2:E3)",
        }
    )
    value_wb = _wb({})
    assert detect_cross_foot_failures(formula_wb, value_wb) == []


def test_reference_in_bounds_allows_unused_range():
    wb = _wb({"A1": 1})
    ref = ParsedReference(raw="Z100", sheet=None, ref="Z100", is_range=False)
    ok, _ = reference_in_bounds(ref, "S", wb)
    assert ok is True


def test_reference_in_bounds_rejects_missing_sheet():
    wb = _wb({"A1": 1})
    ref = ParsedReference(raw="Nope!A1", sheet="Nope", ref="A1", is_range=False)
    ok, reason = reference_in_bounds(ref, "S", wb)
    assert ok is False
    assert "Missing sheet" in reason


def test_reference_in_bounds_rejects_beyond_excel_grid():
    wb = _wb({"A1": 1})
    ref = ParsedReference(raw="A1048577", sheet=None, ref="A1048577", is_range=False)
    ok, reason = reference_in_bounds(ref, "S", wb)
    assert ok is False
    assert "maximum" in reason


def test_year_literals_not_flagged():
    assert extract_numeric_literals("=A1*2025") == []
    assert "1.05" in extract_numeric_literals("=A1*1.05")


def test_check_setting_aliases():
    assert check_setting({"checks": {"fragile_function": "off"}}, "VOLATILE_FUNCTION") == "off"
    assert check_setting({"checks": {"whole_column_reference": "off"}}, "WHOLE_COLUMN_REFERENCE") == "off"
    assert check_setting({"checks": {"cross_foot_failure": "warn"}}, "CROSS_FOOT_FAILURE") == "warn"
    assert check_setting({}, "LIVE_ERROR") == "error"
