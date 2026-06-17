"""Performance guardrails: max_cells / max_formulas / max_reported_findings / timeout."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]


def _make_workbook(tmp_path: Path, rows: int = 30) -> Path:
    """Build a small workbook with `rows` formula rows where each formula
    contains a broken reference to guarantee at least one finding per row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws.append(["Label", "Value", "Formula"])
    for r in range(2, rows + 2):
        ws.cell(row=r, column=1, value=f"Row {r}")
        ws.cell(row=r, column=2, value=r)
        # =SUM(#REF!) triggers BROKEN_REFERENCE on every row.
        ws.cell(row=r, column=3, value="=SUM(#REF!)")
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return path


def _run_audit(workbook: Path, config: dict | None = None, tmp_path: Path | None = None) -> dict:
    config_path = None
    if config is not None:
        assert tmp_path is not None
        config_path = tmp_path / "config.json"
        config_path.write_text(json.dumps(config), encoding="utf-8")
    cmd = [sys.executable, "-m", "spreadsheet_auditor", str(workbook), "--json", "-"]
    if config_path:
        cmd += ["--config", str(config_path)]
    result = subprocess.run(cmd, capture_output=True, text=True)
    assert result.returncode in (0, 1, 2), result.stderr
    return json.loads(result.stdout)


def test_max_formulas_caps_formula_scan(tmp_path):
    workbook = _make_workbook(tmp_path, rows=30)
    payload = _run_audit(workbook, {"limits": {"max_formulas": 5}}, tmp_path)
    assert payload["coverage"]["truncated"]["formulas"] is True
    assert any("Formula scan capped" in note for note in payload["coverage"]["limitations"])


def test_max_reported_findings_marks_truncation(tmp_path):
    workbook = _make_workbook(tmp_path, rows=10)
    payload = _run_audit(workbook, {"limits": {"max_reported_findings": 1}}, tmp_path)
    assert payload["coverage"]["truncated"]["findings"] is True
    assert any("Findings output capped" in note for note in payload["coverage"]["limitations"])


def test_max_cells_marks_truncation(tmp_path):
    workbook = _make_workbook(tmp_path, rows=10)
    payload = _run_audit(workbook, {"limits": {"max_cells": 5}}, tmp_path)
    assert payload["coverage"]["truncated"]["cells"] is True
    assert any("Cell scan capped" in note for note in payload["coverage"]["limitations"])


def test_timeout_zero_is_disabled(tmp_path):
    workbook = _make_workbook(tmp_path, rows=10)
    payload = _run_audit(workbook, {"limits": {"timeout_seconds": 0}}, tmp_path)
    assert payload["coverage"]["truncated"]["timeout"] is False


def test_elapsed_seconds_reported(tmp_path):
    workbook = _make_workbook(tmp_path, rows=10)
    payload = _run_audit(workbook, None, tmp_path)
    assert "elapsed_seconds" in payload["coverage"]
    assert payload["coverage"]["elapsed_seconds"] >= 0.0
