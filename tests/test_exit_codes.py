"""Exit-code semantics: clean audits succeed even with benign limitations."""

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"


def _clean_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Label"
    ws["B1"] = 42
    wb.save(path)


def _no_recalc_config(path: Path) -> None:
    # Disabling recalc deterministically produces a coverage limitation,
    # regardless of whether LibreOffice/defusedxml are installed.
    path.write_text(json.dumps({"recalc": {"enabled": False}}), encoding="utf-8")


def test_clean_workbook_returns_zero_despite_limitations(tmp_path):
    workbook = tmp_path / "clean.xlsx"
    config = tmp_path / "config.json"
    _clean_workbook(workbook)
    _no_recalc_config(config)
    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "audit.py"),
            str(workbook),
            "--config",
            str(config),
            "--json",
            str(tmp_path / "f.json"),
        ],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr


def test_strict_surfaces_limitations(tmp_path):
    workbook = tmp_path / "clean.xlsx"
    config = tmp_path / "config.json"
    _clean_workbook(workbook)
    _no_recalc_config(config)
    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "audit.py"),
            str(workbook),
            "--config",
            str(config),
            "--strict",
            "--json",
            str(tmp_path / "f.json"),
        ],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 2, result.stderr
