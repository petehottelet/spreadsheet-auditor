"""Finance-specific HEUR checks.

These checks are opt-in (off by default). Tests verify both the off-by-default
behavior and the seeded positive cases.
"""

from __future__ import annotations

from openpyxl import Workbook

from spreadsheet_auditor.checks import CheckContext
from spreadsheet_auditor.checks.finance import (
    BalanceSheetBalanceCheck,
    PeriodMismatchCheck,
    SignConventionCheck,
)


def _ctx(wb, config) -> CheckContext:
    allowed = {ws.title for ws in wb.worksheets}
    return CheckContext(
        workbook_path=None,
        formula_wb=wb,
        value_wb=wb,
        allowed_sheet_names=allowed,
        formulas=[],
        config=config,
        inventory={},
    )


def test_balance_sheet_check_off_by_default():
    wb = Workbook()
    ws = wb.active
    ws.title = "BS"
    ws.append(["Total Assets", 100])
    ws.append(["Total Liabilities + Equity", 90])
    assert BalanceSheetBalanceCheck().run(_ctx(wb, {})) == []


def test_balance_sheet_check_fires_when_enabled_and_unbalanced():
    wb = Workbook()
    ws = wb.active
    ws.title = "BS"
    ws.append(["Total Assets", 100])
    ws.append(["Total Liabilities + Equity", 90])
    findings = BalanceSheetBalanceCheck().run(_ctx(wb, {"finance": {"enabled": True}}))
    assert findings, "balance check should fire when assets != L+E"
    assert findings[0].rule_id == "BALANCE_SHEET_BALANCE"
    assert findings[0].detection_mode == "HEUR"
    assert findings[0].error_confidence == "Review"


def test_balance_sheet_check_silent_when_balanced():
    wb = Workbook()
    ws = wb.active
    ws.title = "BS"
    ws.append(["Total Assets", 1000])
    ws.append(["Total Liabilities + Equity", 1000])
    findings = BalanceSheetBalanceCheck().run(_ctx(wb, {"finance": {"enabled": True}}))
    assert findings == []


def test_sign_convention_flags_expense_with_positive_value():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws.append(["Revenue", 100])
    ws.append(["Cost of services", 40])  # positive expense - flag
    findings = SignConventionCheck().run(_ctx(wb, {"finance": {"enabled": True}}))
    assert any(f.rule_id == "SIGN_CONVENTION" for f in findings)


def test_sign_convention_silent_when_expense_negative():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws.append(["Cost of services", -40])
    findings = SignConventionCheck().run(_ctx(wb, {"finance": {"enabled": True}}))
    assert findings == []


def test_period_mismatch_flags_skipped_quarter():
    wb = Workbook()
    ws = wb.active
    ws.title = "Q"
    ws.append(["Line", "Q1 2024", "Q3 2024", "Q4 2024"])  # missing Q2
    findings = PeriodMismatchCheck().run(_ctx(wb, {"finance": {"enabled": True}}))
    assert any(f.rule_id == "PERIOD_MISMATCH" for f in findings)


def test_period_mismatch_silent_with_proper_sequence():
    wb = Workbook()
    ws = wb.active
    ws.title = "Q"
    ws.append(["Line", "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"])
    findings = PeriodMismatchCheck().run(_ctx(wb, {"finance": {"enabled": True}}))
    assert findings == []
