"""Generate the demo workbook plus its audit outputs.

Builds `examples/demo_bad_budget.xlsx` with a curated set of seeded defects that
exercise every deterministic rule the auditor advertises, then runs the auditor
to refresh `demo_audit_report.md`, `demo_findings.json`, `demo_audit_report.html`,
and `demo_annotated.xlsx`.

Run from the repo root:

    python examples/make_demo.py
"""

from __future__ import annotations

import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
EXAMPLES = ROOT / "examples"
WORKBOOK = EXAMPLES / "demo_bad_budget.xlsx"
REPORT_MD = EXAMPLES / "demo_audit_report.md"
REPORT_HTML = EXAMPLES / "demo_audit_report.html"
FINDINGS_JSON = EXAMPLES / "demo_findings.json"
ANNOTATED = EXAMPLES / "demo_annotated.xlsx"
# Packaged copy that ships inside the wheel/skill so `--demo` works when the
# repo `examples/` directory is not on disk. Kept in sync from WORKBOOK below.
PACKAGED_DEMO = ROOT / "spreadsheet_auditor" / "demo" / "demo_bad_budget.xlsx"


def build_workbook() -> None:
    """Build a small budget workbook seeded with realistic-looking defects.

    The workbook is intentionally readable: rows are labelled so reviewers can
    see why each finding is interesting, and the seeded defects map 1:1 to the
    PRD's acceptance criteria for the demo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"

    # Header
    ws.append(["Line item", "Jan", "Feb", "Mar", "Q1 total"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    # Revenue & expense block. Note row 5 is the "missing" item that the off-by-one
    # SUM should have included; the formula drift in F6 omits column D.
    ws.append(["Subscriptions", 1200, 1320, 1450, "=SUM(B2:D2)"])  # row 2
    ws.append(["Services", 800, 820, 900, "=SUM(B3:D3)"])  # row 3
    ws.append(["Subtotal", "=SUM(B2:B3)", "=SUM(C2:C3)", "=SUM(D2:D3)", "=SUM(B4:D4)"])  # row 4
    # Off-by-one: total for "Other revenue" lives below, but the grand total
    # at B11 stops at B9 instead of B10.
    ws.append(["Other revenue", 150, 175, 200, "=SUM(B5:D5)"])  # row 5
    # Formula drift: the grand total row should be =SUM(*2:*5), but F6 ranges
    # over B6:D6 even though there is no data in column D for row 6.
    ws.append(["Revenue total", "=SUM(B2:B5)+B5", "=SUM(C2:C5)", "=SUM(D2:D5)", "=SUM(B6:C6)"])  # row 6 drift

    # COGS block with a hardcoded plug in the formula column.
    ws.append(["Hosting", 200, 210, 220, "=SUM(B7:D7)"])  # row 7
    ws.append(["Support staff", 600, 620, 640, "=SUM(B8:D8)"])  # row 8
    ws.append(["Marketing", 300, 320, 340, 999])  # row 9 hardcoded plug in formula block
    ws.append(["COGS total", "=SUM(B7:B9)", "=SUM(C7:C9)", "=SUM(D7:D9)", "=SUM(B10:D10)"])  # row 10
    # Off-by-one grand total: should run B2:B9 across revenue rows but stops at B8,
    # silently dropping Marketing (row 9).
    ws.append(["Grand total", "=SUM(B2:B8)", "=SUM(C2:C9)", "=SUM(D2:D9)", "=SUM(B11:D11)"])  # row 11

    # Hidden row that still feeds an aggregate.
    ws.append(["Adjustment (hidden)", 50, 60, 70, "=SUM(B12:D12)"])  # row 12
    ws.row_dimensions[12].hidden = True
    ws.append(["With adjustment", "=B11+B12", "=C11+C12", "=D11+D12", "=SUM(B13:D13)"])  # row 13

    # Live #REF! and broken reference.
    ws.append(["Stale link", "=SUM(#REF!)", None, None, None])  # row 14

    # Lookup table (duplicate keys + whitespace).
    ws.append([])  # row 15
    ws.append(["Region", "Owner", None, None, None])  # row 16
    ws.append(["North", "Alex", None, None, None])  # row 17
    ws.append([" North ", "Sam", None, None, None])  # row 18: whitespace + duplicate
    ws.append(["South", "Jess", None, None, None])  # row 19

    # Number stored as text in a numeric column.
    ws.append(["Bonus pool", "1,250", None, None, None])  # row 20

    # Reference sheet for cross-sheet defect.
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Headline KPI"
    summary["B1"] = "Total revenue"
    summary["C1"] = "=Budget!B6"  # depends on the drift'd row 6

    wb.save(WORKBOOK)
    PACKAGED_DEMO.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(WORKBOOK, PACKAGED_DEMO)


def regenerate_outputs() -> int:
    cmd = [
        sys.executable,
        "-m",
        "spreadsheet_auditor",
        str(WORKBOOK),
        "--out",
        str(REPORT_MD),
        "--json",
        str(FINDINGS_JSON),
        "--annotated",
        str(ANNOTATED),
    ]
    # Auditor exits 1 when findings are present; that is the expected outcome
    # here because the demo workbook is intentionally broken.
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode not in (0, 1):
        sys.stderr.write(result.stderr)
        return result.returncode

    # Best-effort HTML rendering (added once the HTML output milestone lands).
    html_cmd = [
        sys.executable,
        "-m",
        "spreadsheet_auditor",
        str(WORKBOOK),
        "--out",
        str(REPORT_HTML),
        "--format",
        "html",
    ]
    html_result = subprocess.run(html_cmd, capture_output=True, text=True)
    if html_result.returncode not in (0, 1):
        # HTML output may not be wired yet; do not fail demo generation.
        pass
    return 0


def main() -> int:
    EXAMPLES.mkdir(parents=True, exist_ok=True)
    build_workbook()
    print(f"Wrote {WORKBOOK}")
    print(f"Synced packaged demo copy {PACKAGED_DEMO}")
    code = regenerate_outputs()
    if code == 0:
        for path in [REPORT_MD, FINDINGS_JSON, ANNOTATED, REPORT_HTML]:
            if path.exists():
                print(f"Wrote {path}")
    return code


if __name__ == "__main__":
    raise SystemExit(main())
