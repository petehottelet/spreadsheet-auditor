from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment


def _anchor_location(location: str) -> tuple[str, str] | None:
    """Resolve a finding location to a single (sheet, coordinate) anchor cell.

    Handles comma-joined multi-locations (uses the first) and range locations
    (uses the top-left cell), e.g. "Sheet1!A1:B2" -> ("Sheet1", "A1").
    """
    if not location:
        return None
    first = location.split(",", 1)[0].strip()
    if "!" not in first:
        return None
    sheet_name, coord = first.split("!", 1)
    coord = coord.split(":", 1)[0].replace("$", "").strip()
    if not sheet_name or not coord:
        return None
    return sheet_name, coord


def annotate_workbook(source_path: str | Path, output_path: str | Path, findings: list[dict]) -> None:
    source = Path(source_path)
    keep_vba = source.suffix.lower() == ".xlsm"
    wb = load_workbook(source, keep_vba=keep_vba)
    for finding in findings:
        if finding.get("suppressed"):
            continue
        location = finding.get("location", "")
        anchor = _anchor_location(location)
        if anchor is None:
            continue
        sheet_name, coord = anchor
        if sheet_name not in wb.sheetnames:
            continue
        cell = wb[sheet_name][coord]
        text = (
            f"{finding.get('severity')} {finding.get('rule_id')}\n"
            f"{finding.get('title')}\n"
            f"Fix: {finding.get('suggested_fix')}"
        )
        cell.comment = Comment(text, "Spreadsheet Auditor")
    wb.save(output_path)
