from __future__ import annotations

from collections.abc import Iterable

from openpyxl.utils.cell import get_column_letter, range_boundaries

from formula_parser import ParsedReference


def ref_location(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def range_size(ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return (max_col - min_col + 1) * (max_row - min_row + 1)


def expand_reference(ref: ParsedReference, default_sheet: str, limit: int = 1000) -> list[str]:
    sheet = ref.sheet or default_sheet
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref.ref)
    except ValueError:
        return []
    if (max_col - min_col + 1) * (max_row - min_row + 1) > limit:
        return []
    cells: list[str] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cells.append(ref_location(sheet, row, col))
    return cells


def reference_in_bounds(ref: ParsedReference, default_sheet: str, workbook) -> tuple[bool, str | None]:
    sheet = ref.sheet or default_sheet
    if sheet not in workbook.sheetnames:
        return False, f"Missing sheet '{sheet}'"
    ws = workbook[sheet]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref.ref)
    except ValueError:
        return False, "Unsupported reference syntax"
    if max_row > ws.max_row or max_col > ws.max_column:
        return False, f"{sheet}!{ref.ref} is outside the used range {ws.max_row}x{ws.max_column}"
    return True, None


def cells_from_locations(workbook, locations: Iterable[str]):
    for location in locations:
        if "!" not in location:
            continue
        sheet, coord = location.split("!", 1)
        if sheet in workbook.sheetnames:
            yield location, workbook[sheet][coord]
