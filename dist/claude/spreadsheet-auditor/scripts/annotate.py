from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment


def annotate_workbook(source_path: str | Path, output_path: str | Path, findings: list[dict]) -> None:
    source = Path(source_path)
    keep_vba = source.suffix.lower() == ".xlsm"
    wb = load_workbook(source, keep_vba=keep_vba)
    for finding in findings:
        if finding.get("suppressed"):
            continue
        location = finding.get("location", "")
        if "!" not in location or "," in location or ":" in location:
            continue
        sheet_name, coord = location.split("!", 1)
        if sheet_name not in wb.sheetnames:
            continue
        cell = wb[sheet_name][coord]
        text = (
            f"{finding.get('severity')} {finding.get('rule_id')}\n"
            f"{finding.get('title')}\n"
            f"Fix: {finding.get('suggested_fix')}"
        )
        cell.comment = Comment(text, "Spreadsheet Auditor")
    wb.save(output_path)
