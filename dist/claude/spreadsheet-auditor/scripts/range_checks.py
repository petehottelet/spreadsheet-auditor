from __future__ import annotations

import re
from typing import Iterable

from openpyxl.utils.cell import get_column_letter, range_boundaries

from finding import Finding
from formula_parser import extract_functions, extract_references, is_formula
from workbook_inventory import location


AGG_FUNCS = {"SUM", "AVERAGE", "COUNT", "COUNTA"}
SUBTOTAL_WORDS = {"total", "subtotal", "grand total", "sum"}


def aggregate_ranges(formula: str):
    funcs = extract_functions(formula)
    if not funcs.intersection(AGG_FUNCS):
        return []
    return [ref for ref in extract_references(formula) if ref.is_range]


def detect_range_issues(formula_wb, value_wb, formula_cells: list[dict]) -> list[Finding]:
    findings: list[Finding] = []
    for cell in formula_cells:
        for ref in aggregate_ranges(cell["formula"]):
            sheet = ref.sheet or cell["sheet"]
            if sheet not in formula_wb.sheetnames:
                continue
            ws = formula_wb[sheet]
            try:
                min_col, min_row, max_col, max_row = range_boundaries(ref.ref)
            except ValueError:
                continue
            findings.extend(_detect_exclusion(ws, cell, ref.ref, min_col, min_row, max_col, max_row))
            findings.extend(_detect_subtotal_inclusion(ws, cell, ref.ref, min_col, min_row, max_col, max_row))
            findings.extend(_detect_hidden_intersections(ws, cell, ref.ref, min_col, min_row, max_col, max_row))
    return findings


def _cell_has_data(cell) -> bool:
    value = cell.value
    if value is None:
        return False
    if is_formula(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def _row_label(ws, row: int, min_col: int) -> str:
    labels = []
    for col in range(1, min_col):
        value = ws.cell(row=row, column=col).value
        if value is not None:
            labels.append(str(value).strip())
    return " ".join(labels).lower()


def _detect_exclusion(ws, formula_cell: dict, ref_text: str, min_col: int, min_row: int, max_col: int, max_row: int):
    findings: list[Finding] = []
    loc = formula_cell["location"]
    formula = formula_cell["formula"]

    if min_col == max_col:
        candidates = []
        if min_row > 1:
            candidates.append((min_row - 1, min_col, "above"))
        if max_row + 1 < formula_cell["row"]:
            candidates.append((max_row + 1, min_col, "below"))
        for row, col, direction in candidates:
            neighbor = ws.cell(row=row, column=col)
            if _cell_has_data(neighbor):
                label = _row_label(ws, row, min_col)
                if any(word in label for word in SUBTOTAL_WORDS):
                    continue
                findings.append(
                    Finding(
                        rule_id="RANGE_EXCLUSION",
                        severity="Critical",
                        error_confidence="Likely defect",
                        detection_mode="DET",
                        location=loc,
                        title="Aggregation range appears to exclude adjacent data row",
                        formula=formula,
                        evidence=[f"{ws.title}!{ref_text} excludes adjacent {direction} cell {ws.title}!{neighbor.coordinate} with value {neighbor.value!r}."],
                        suggested_fix=f"Confirm whether {ws.title}!{neighbor.coordinate} belongs in the aggregate, then extend the range if appropriate.",
                    )
                )
    if min_row == max_row:
        candidates = []
        if min_col > 1:
            candidates.append((min_row, min_col - 1, "left"))
        if max_col + 1 < formula_cell["col"]:
            candidates.append((min_row, max_col + 1, "right"))
        for row, col, direction in candidates:
            neighbor = ws.cell(row=row, column=col)
            if _cell_has_data(neighbor):
                findings.append(
                    Finding(
                        rule_id="RANGE_EXCLUSION",
                        severity="Critical",
                        error_confidence="Likely defect",
                        detection_mode="DET",
                        location=loc,
                        title="Aggregation range appears to exclude adjacent data column",
                        formula=formula,
                        evidence=[f"{ws.title}!{ref_text} excludes adjacent {direction} cell {ws.title}!{neighbor.coordinate} with value {neighbor.value!r}."],
                        suggested_fix=f"Confirm whether {ws.title}!{neighbor.coordinate} belongs in the aggregate, then extend the range if appropriate.",
                    )
                )
    return findings


def _detect_subtotal_inclusion(ws, formula_cell: dict, ref_text: str, min_col: int, min_row: int, max_col: int, max_row: int):
    findings: list[Finding] = []
    for row in range(min_row, max_row + 1):
        label = _row_label(ws, row, min_col)
        if any(word in label for word in SUBTOTAL_WORDS):
            findings.append(
                Finding(
                    rule_id="RANGE_INCLUDES_SUBTOTAL",
                    severity="High",
                    error_confidence="Likely defect",
                    detection_mode="DET",
                    location=formula_cell["location"],
                    title="Aggregation range appears to include subtotal or total row",
                    formula=formula_cell["formula"],
                    evidence=[f"{ws.title}!{ref_text} includes row {row}, labeled {label!r}."],
                    suggested_fix="Review the aggregate range and exclude subtotal/total rows unless intentionally double-counting.",
                )
            )
    return findings


def _detect_hidden_intersections(ws, formula_cell: dict, ref_text: str, min_col: int, min_row: int, max_col: int, max_row: int):
    hidden_rows = [row for row in range(min_row, max_row + 1) if ws.row_dimensions[row].hidden]
    hidden_cols = [
        get_column_letter(col)
        for col in range(min_col, max_col + 1)
        if ws.column_dimensions[get_column_letter(col)].hidden
    ]
    if not hidden_rows and not hidden_cols:
        return []
    pieces = []
    if hidden_rows:
        pieces.append(f"hidden rows {hidden_rows}")
    if hidden_cols:
        pieces.append(f"hidden columns {hidden_cols}")
    return [
        Finding(
            rule_id="HIDDEN_STRUCTURE_IN_TOTAL",
            severity="Medium",
            error_confidence="Review",
            detection_mode="DET",
            location=formula_cell["location"],
            title="Aggregate includes hidden structure",
            formula=formula_cell["formula"],
            evidence=[f"{ws.title}!{ref_text} intersects " + ", ".join(pieces) + "."],
            suggested_fix="Confirm hidden inputs are intentional and disclosed in visible workbook documentation.",
        )
    ]


def detect_literal_constants(formula_cells: list[dict]) -> list[Finding]:
    from formula_parser import extract_numeric_literals

    findings: list[Finding] = []
    for cell in formula_cells:
        literals = extract_numeric_literals(cell["formula"])
        if literals:
            findings.append(
                Finding(
                    rule_id="LITERAL_CONSTANT",
                    severity="Medium",
                    error_confidence="Review",
                    detection_mode="DET",
                    location=cell["location"],
                    title="Formula contains embedded numeric literal",
                    formula=cell["formula"],
                    evidence=[f"Non-trivial numeric literal(s) found: {', '.join(literals)}."],
                    suggested_fix="Move the assumption to a labeled input cell and reference it from the formula.",
                )
            )
    return findings


def detect_fragile_functions(formula_cells: list[dict]) -> list[Finding]:
    findings: list[Finding] = []
    volatile = {"OFFSET", "INDIRECT", "NOW", "RAND", "RANDBETWEEN", "TODAY"}
    whole_col_re = re.compile(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}")
    for cell in formula_cells:
        funcs = extract_functions(cell["formula"])
        if funcs.intersection(volatile):
            findings.append(
                Finding(
                    rule_id="VOLATILE_FUNCTION",
                    severity="Medium",
                    error_confidence="Review",
                    detection_mode="DET",
                    location=cell["location"],
                    title="Formula uses volatile or fragile function",
                    formula=cell["formula"],
                    evidence=[f"Function(s) found: {', '.join(sorted(funcs.intersection(volatile)))}."],
                    suggested_fix="Confirm the volatility is intentional; prefer stable bounded references when possible.",
                )
            )
        if whole_col_re.search(cell["formula"].upper()):
            findings.append(
                Finding(
                    rule_id="WHOLE_COLUMN_REFERENCE",
                    severity="Medium",
                    error_confidence="Review",
                    detection_mode="DET",
                    location=cell["location"],
                    title="Formula references an entire column",
                    formula=cell["formula"],
                    evidence=["Whole-column references can hide range and performance issues."],
                    suggested_fix="Use a bounded range sized to the actual table.",
                )
            )
    return findings
