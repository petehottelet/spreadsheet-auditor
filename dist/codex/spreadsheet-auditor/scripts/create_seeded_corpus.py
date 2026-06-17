from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CORPUS = ROOT / "assets" / "test_corpus"


def build_seeded_defects() -> None:
    out_dir = CORPUS / "seeded_defects"
    expected_dir = CORPUS / "expected_findings"
    out_dir.mkdir(parents=True, exist_ok=True)
    expected_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Line"
    ws["B1"] = "Y1"
    ws["C1"] = "Y2"
    ws["D1"] = "Y3"
    ws["E1"] = "Total"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    ws.append(["Revenue", 100, 120, 130, "=SUM(B2:D2)"])
    ws.append(["Expense", -40, -45, -50, "=SUM(B3:D3)"])
    ws.append(["EBITDA", "=B2+B3", "=C2+C3", "=D2*D3", "=SUM(B4:D4)"])
    ws.append(["Margin", "=B4/B2", "=C4/C2", "=D4/D2", "=E4/E2"])

    ws["A7"] = "Range exclusion block"
    ws["A8"], ws["B8"] = "Item A", 10
    ws["A9"], ws["B9"] = "Item B", 20
    ws["A10"], ws["B10"] = "Item C omitted", 30
    ws["A11"], ws["B11"] = "Total", "=SUM(B8:B9)"

    ws["A14"] = "Subtotal inclusion block"
    ws["A15"], ws["B15"] = "Component A", 5
    ws["A16"], ws["B16"] = "Subtotal", "=SUM(B15:B15)"
    ws["A17"], ws["B17"] = "Component B", 7
    ws["A18"], ws["B18"] = "Grand total includes subtotal", "=SUM(B15:B17)"

    ws["A21"] = "Hardcode in formula block"
    ws["B22"] = "=B2+B3"
    ws["C22"] = 999
    ws["D22"] = "=D2+D3"

    ws["A25"], ws["B25"] = "Embedded literal", "=B2*1.05"
    ws["A27"], ws["B27"] = "IFERROR mask", '=IFERROR(B2/B99,"")'
    ws["A28"], ws["B28"] = "Broken ref", "=SUM(#REF!)"

    ws["A30"], ws["B30"] = "Visible input", 5
    ws["A31"], ws["B31"] = "Hidden input", 6
    ws.row_dimensions[31].hidden = True
    ws["A32"], ws["B32"] = "Total includes hidden row", "=SUM(B30:B31)"

    ws["A35"], ws["B35"] = "Text number", "1,234"
    ws["A36"] = " Customer A "
    ws["A37"] = "SKU1"
    ws["A38"] = " sku1 "
    ws.merge_cells("A42:B42")
    ws["A42"] = "Merged note"
    ws["A44"], ws["B44"] = "Live error value", "#VALUE!"

    out_file = out_dir / "seeded-defects.xlsx"
    wb.save(out_file)

    expected = {
        "workbook": "seeded-defects.xlsx",
        "expected_rule_ids": [
            "LIVE_ERROR",
            "BROKEN_REFERENCE",
            "FORMULA_DRIFT",
            "RANGE_EXCLUSION",
            "RANGE_INCLUDES_SUBTOTAL",
            "HARDCODE_IN_FORMULA_BLOCK",
            "LITERAL_CONSTANT",
            "IFERROR_MASK",
            "HIDDEN_STRUCTURE_IN_TOTAL",
            "NUMBERS_STORED_AS_TEXT",
            "WHITESPACE_KEY",
            "DUPLICATE_KEY",
            "MERGED_CELL_IN_DATA_RANGE"
        ]
    }
    (expected_dir / "seeded-defects.json").write_text(json.dumps(expected, indent=2), encoding="utf-8")
    print(out_file)


if __name__ == "__main__":
    build_seeded_defects()
