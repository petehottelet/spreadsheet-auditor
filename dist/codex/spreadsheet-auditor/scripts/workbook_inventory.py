from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from formula_parser import is_formula


ERROR_VALUES = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA"}


def load_workbooks(path: str | Path):
    keep_vba = Path(path).suffix.lower() == ".xlsm"
    formula_wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
    value_wb = load_workbook(path, data_only=True, keep_vba=keep_vba)
    return formula_wb, value_wb


def location(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def iter_used_cells(workbook):
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    yield ws, cell


def formula_cells(workbook) -> list[dict[str, Any]]:
    cells = []
    for ws, cell in iter_used_cells(workbook):
        if is_formula(cell.value):
            cells.append(
                {
                    "sheet": ws.title,
                    "row": cell.row,
                    "col": cell.column,
                    "coord": cell.coordinate,
                    "location": location(ws.title, cell.row, cell.column),
                    "formula": str(cell.value),
                }
            )
    return cells


def inventory(path: str | Path, formula_wb, value_wb, preflight: dict) -> dict[str, Any]:
    formulas = formula_cells(formula_wb)
    sheets = []
    for ws in formula_wb.worksheets:
        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
        hidden_cols = [col for col, dim in ws.column_dimensions.items() if dim.hidden]
        sheets.append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "used_range": f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                "hidden_rows": hidden_rows,
                "hidden_columns": hidden_cols,
                "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            }
        )

    external_links = []
    for link in getattr(formula_wb, "_external_links", []) or []:
        external_links.append(str(getattr(link, "file_link", link)))

    return {
        "path": str(path),
        "sha256": preflight["sha256"],
        "extension": preflight["extension"],
        "sheets": sheets,
        "sheets_analyzed": len(sheets),
        "formulas_scanned": len(formulas),
        "external_links": external_links,
        "macros_present": bool(preflight.get("macros_present")),
        "macros_executed": False,
    }


def scan_live_errors(formula_wb, value_wb):
    findings = []
    for wb in (value_wb, formula_wb):
        for ws, cell in iter_used_cells(wb):
            if isinstance(cell.value, str) and cell.value in ERROR_VALUES:
                findings.append((location(ws.title, cell.row, cell.column), cell.value))
    return sorted(set(findings))
