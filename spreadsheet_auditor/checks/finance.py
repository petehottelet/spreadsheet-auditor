"""Opt-in finance-specific HEUR checks.

These checks are off by default. Enable them with config:

    {
      "finance": { "enabled": true },
      "checks": {
        "BALANCE_SHEET_BALANCE": "warn",
        "SIGN_CONVENTION":       "warn",
        "PERIOD_MISMATCH":       "warn"
      }
    }

They are intentionally label-driven (rather than ontology-driven) so they stay
predictable and easy to override per-project. All findings default to a `Review`
confidence: surface for inspection, never claim certainty.
"""

from __future__ import annotations

import re
from typing import Iterable

from ..finding import Finding
from .base import Check, CheckContext, register


_BS_ASSET_LABELS = re.compile(r"\btotal\s+assets\b", re.IGNORECASE)
_BS_LIAB_LABELS = re.compile(r"\btotal\s+liab(ilit(ies|y))?\s*(\+|and|&)?\s*(equity)?\b", re.IGNORECASE)
_EXPENSE_LABELS = re.compile(r"\b(expense|cost|outflow|cogs|opex|capex|tax(?:es)?)\b", re.IGNORECASE)
_REVENUE_LABELS = re.compile(r"\b(revenue|sales|inflow|income)\b", re.IGNORECASE)
_PERIOD_HEADER = re.compile(r"\b(Q[1-4])\s*(?:[-/]?\s*)?(?:FY)?\s*(\d{2,4})\b")


def _finance_enabled(config: dict) -> bool:
    return bool((config.get("finance") or {}).get("enabled"))


def _iter_label_rows(formula_wb, allowed_sheet_names: set[str]) -> Iterable[tuple[str, int, str, list]]:
    for ws in formula_wb.worksheets:
        if ws.title not in allowed_sheet_names:
            continue
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            if not isinstance(label, str):
                continue
            row_values = [
                ws.cell(row=row, column=col).value
                for col in range(2, ws.max_column + 1)
            ]
            yield ws.title, row, label, row_values


@register
class BalanceSheetBalanceCheck(Check):
    name = "balance_sheet_balance"
    description = "Flags rows labelled 'Total Assets' and 'Total Liabilities + Equity' whose totals disagree."
    rule_ids = ("BALANCE_SHEET_BALANCE",)
    mode = "HEUR"

    def run(self, ctx: CheckContext) -> list[Finding]:
        if not _finance_enabled(ctx.config):
            return []
        assets: dict[tuple[str, int], list] = {}
        liab_equity: dict[tuple[str, int], list] = {}
        for sheet, row, label, values in _iter_label_rows(ctx.formula_wb, ctx.allowed_sheet_names):
            if _BS_ASSET_LABELS.search(label):
                assets[(sheet, row)] = values
            elif _BS_LIAB_LABELS.search(label):
                liab_equity[(sheet, row)] = values

        findings: list[Finding] = []
        for (sheet, asset_row), asset_values in assets.items():
            # Same sheet, look for nearest liab+equity row.
            sheet_liab = [(r, vals) for (s, r), vals in liab_equity.items() if s == sheet]
            if not sheet_liab:
                continue
            liab_row, liab_values = min(sheet_liab, key=lambda item: abs(item[0] - asset_row))
            for col_idx, (a, l) in enumerate(zip(asset_values, liab_values), start=2):
                if isinstance(a, (int, float)) and isinstance(l, (int, float)):
                    diff = abs(a - l)
                    threshold = max(abs(a), abs(l)) * 0.005 + 1e-9
                    if diff > threshold and diff >= 1:
                        from openpyxl.utils import get_column_letter

                        findings.append(
                            Finding(
                                rule_id="BALANCE_SHEET_BALANCE",
                                severity="Medium",
                                error_confidence="Review",
                                detection_mode="HEUR",
                                location=f"{sheet}!{get_column_letter(col_idx)}{asset_row}",
                                title="Balance sheet does not balance",
                                evidence=[
                                    f"Total Assets ({a}) at row {asset_row} disagrees with Total Liabilities+Equity ({l}) at row {liab_row} by {diff}.",
                                ],
                                suggested_fix="Reconcile the asset and liability+equity totals. Common causes: missing line item, sign error, missing reclassification.",
                            )
                        )
        return findings


@register
class SignConventionCheck(Check):
    name = "sign_convention"
    description = "Flags rows labelled as expenses/costs that contain positive numbers (or revenue with negatives)."
    rule_ids = ("SIGN_CONVENTION",)
    mode = "HEUR"

    def run(self, ctx: CheckContext) -> list[Finding]:
        if not _finance_enabled(ctx.config):
            return []
        findings: list[Finding] = []
        for sheet, row, label, values in _iter_label_rows(ctx.formula_wb, ctx.allowed_sheet_names):
            label_clean = label.strip()
            is_expense = bool(_EXPENSE_LABELS.search(label_clean))
            is_revenue = bool(_REVENUE_LABELS.search(label_clean)) and not is_expense
            if not (is_expense or is_revenue):
                continue
            for col_offset, value in enumerate(values, start=2):
                if not isinstance(value, (int, float)):
                    continue
                if is_expense and value > 0 and value > 1e-6:
                    findings.append(
                        _sign_finding(sheet, row, col_offset, label_clean, value, kind="expense_positive")
                    )
                elif is_revenue and value < 0 and value < -1e-6:
                    findings.append(
                        _sign_finding(sheet, row, col_offset, label_clean, value, kind="revenue_negative")
                    )
        return findings


def _sign_finding(sheet: str, row: int, col_offset: int, label: str, value, kind: str) -> Finding:
    from openpyxl.utils import get_column_letter

    expectation = "negative" if kind == "expense_positive" else "positive"
    return Finding(
        rule_id="SIGN_CONVENTION",
        severity="Medium",
        error_confidence="Review",
        detection_mode="HEUR",
        location=f"{sheet}!{get_column_letter(col_offset)}{row}",
        title="Row label sign convention mismatch",
        evidence=[
            f"Row labelled {label!r} has value {value}, which is unexpected for the {kind.split('_')[0]} sign convention (expected {expectation}).",
        ],
        suggested_fix="Confirm the sign convention for this row. Many models keep expenses negative throughout and gross up only at the EBITDA line.",
    )


@register
class PeriodMismatchCheck(Check):
    name = "period_mismatch"
    description = "Flags period headers like 'Q1 2024', 'Q3 2024' (missing Q2) on the same row."
    rule_ids = ("PERIOD_MISMATCH",)
    mode = "HEUR"

    def run(self, ctx: CheckContext) -> list[Finding]:
        if not _finance_enabled(ctx.config):
            return []
        findings: list[Finding] = []
        from openpyxl.utils import get_column_letter

        for ws in ctx.formula_wb.worksheets:
            if ws.title not in ctx.allowed_sheet_names:
                continue
            for row in range(1, min(6, ws.max_row + 1)):  # header row scan
                headers = [
                    (col, ws.cell(row=row, column=col).value)
                    for col in range(1, ws.max_column + 1)
                ]
                parsed = []
                for col, value in headers:
                    if not isinstance(value, str):
                        continue
                    match = _PERIOD_HEADER.search(value)
                    if match:
                        parsed.append((col, match.group(1).upper(), match.group(2)))
                if len(parsed) < 2:
                    continue
                # If quarters appear in the same row but skip one, flag the gap.
                quarters = [(col, int(q[1]), year) for col, q, year in parsed]
                quarters.sort(key=lambda item: item[0])
                for prev, curr in zip(quarters, quarters[1:]):
                    expected = ((prev[1]) % 4) + 1
                    if prev[2] == curr[2] and curr[1] != expected:
                        findings.append(
                            Finding(
                                rule_id="PERIOD_MISMATCH",
                                severity="Medium",
                                error_confidence="Review",
                                detection_mode="HEUR",
                                location=f"{ws.title}!{get_column_letter(curr[0])}{row}",
                                title="Period header skips a quarter",
                                evidence=[
                                    f"Header at column {get_column_letter(prev[0])} is Q{prev[1]} {prev[2]} but the next header is Q{curr[1]} {curr[2]}.",
                                ],
                                suggested_fix="Verify column ordering. A typical sequence is Q1 -> Q2 -> Q3 -> Q4.",
                            )
                        )
        return findings
