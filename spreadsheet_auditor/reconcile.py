from __future__ import annotations

from openpyxl.utils.cell import get_column_letter, range_boundaries

from .finding import Finding
from .formula_parser import extract_functions, is_formula
from .range_checks import AGG_FUNCS, aggregate_ranges


def _numeric(value) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _is_aggregate_formula(value) -> bool:
    return is_formula(value) and bool(extract_functions(str(value)).intersection(AGG_FUNCS))


def detect_total_mismatches(formula_wb, value_wb, formula_cells: list[dict], tolerance: float = 1e-6) -> list[Finding]:
    findings: list[Finding] = []
    for cell in formula_cells:
        ranges = aggregate_ranges(cell["formula"])
        if len(ranges) != 1:
            continue
        ref = ranges[0]
        sheet = ref.sheet or cell["sheet"]
        if sheet not in value_wb.sheetnames:
            continue
        value_ws = value_wb[sheet]
        formula_value_ws = value_wb[cell["sheet"]]
        cached_total = _numeric(formula_value_ws[cell["coord"]].value)
        if cached_total is None:
            continue
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref.ref)
        except ValueError:
            continue
        component_sum = 0.0
        numeric_count = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = _numeric(value_ws.cell(row=row, column=col).value)
                if value is not None:
                    component_sum += value
                    numeric_count += 1
        if numeric_count and abs(cached_total - component_sum) > tolerance:
            findings.append(
                Finding(
                    rule_id="TOTAL_MISMATCH",
                    severity="Critical",
                    error_confidence="Defect",
                    detection_mode="DET",
                    location=cell["location"],
                    title="Stated total differs from referenced components",
                    formula=cell["formula"],
                    evidence=[f"Cached value is {cached_total}; recomputed referenced components sum to {component_sum}."],
                    impact={"estimated_delta": cached_total - component_sum},
                    suggested_fix="Recalculate the workbook and review the aggregate formula and referenced component range.",
                )
            )
    return findings


def detect_cross_foot_failures(
    formula_wb,
    value_wb,
    allowed_sheet_names: set[str] | None = None,
    tolerance: float = 1e-6,
) -> list[Finding]:
    """Compare a grand total reached down a totals column vs across a totals row.

    Scoped to explicit rectangular tables: the grand-total corner must have a
    vertical run of >=2 aggregate formulas directly above it (row totals) and a
    horizontal run of >=2 aggregate formulas directly to its left (column
    totals). Value-gated: skips silently when cached/recalculated values are
    unavailable rather than emitting false positives.
    """
    findings: list[Finding] = []
    for ws in formula_wb.worksheets:
        if allowed_sheet_names is not None and ws.title not in allowed_sheet_names:
            continue
        if ws.title not in value_wb.sheetnames:
            continue
        value_ws = value_wb[ws.title]
        for gr in range(2, ws.max_row + 1):
            for gc in range(2, ws.max_column + 1):
                rows: list[int] = []
                r = gr - 1
                while r >= 1 and _is_aggregate_formula(ws.cell(row=r, column=gc).value):
                    rows.append(r)
                    r -= 1
                if len(rows) < 2:
                    continue
                cols: list[int] = []
                c = gc - 1
                while c >= 1 and _is_aggregate_formula(ws.cell(row=gr, column=c).value):
                    cols.append(c)
                    c -= 1
                if len(cols) < 2:
                    continue

                down = [v for v in (_numeric(value_ws.cell(row=rr, column=gc).value) for rr in rows) if v is not None]
                across = [v for v in (_numeric(value_ws.cell(row=gr, column=cc).value) for cc in cols) if v is not None]
                if len(down) < 2 or len(across) < 2:
                    continue

                down_sum = sum(down)
                across_sum = sum(across)
                if abs(down_sum - across_sum) > tolerance:
                    corner = f"{ws.title}!{get_column_letter(gc)}{gr}"
                    findings.append(
                        Finding(
                            rule_id="CROSS_FOOT_FAILURE",
                            severity="Critical",
                            error_confidence="Defect",
                            detection_mode="DET",
                            location=corner,
                            title="Row totals and column totals disagree",
                            evidence=[
                                f"Sum of row totals down column {get_column_letter(gc)} is {down_sum}; "
                                f"sum of column totals across row {gr} is {across_sum}."
                            ],
                            impact={"estimated_delta": down_sum - across_sum},
                            suggested_fix="Reconcile the totals row and totals column; one of the contributing aggregates is likely wrong.",
                        )
                    )
    return findings
