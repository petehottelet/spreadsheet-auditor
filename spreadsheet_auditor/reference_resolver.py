from __future__ import annotations

from collections.abc import Iterable

from openpyxl.utils.cell import get_column_letter, range_boundaries

from .formula_parser import ParsedReference

EXCEL_MAX_ROW = 1048576
EXCEL_MAX_COL = 16384


def ref_location(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def range_size(ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return (max_col - min_col + 1) * (max_row - min_row + 1)


def expand_reference(ref: ParsedReference, default_sheet: str, limit: int = 1000) -> list[str]:
    sheet = ref.sheet or default_sheet
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref.ref)
    except ValueError:
        return []
    if (max_col - min_col + 1) * (max_row - min_row + 1) > limit:
        return []
    cells: list[str] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cells.append(ref_location(sheet, row, col))
    return cells


def reference_in_bounds(ref: ParsedReference, default_sheet: str, workbook) -> tuple[bool, str | None]:
    sheet = ref.sheet or default_sheet
    if sheet not in workbook.sheetnames:
        return False, f"Missing sheet '{sheet}'"
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref.ref)
    except ValueError:
        return False, "Unsupported reference syntax"
    # Referencing cells beyond the populated/used range is normal Excel usage
    # (intentionally oversized ranges, future inputs). Only flag references that
    # fall outside Excel's absolute grid.
    if max_row is not None and max_row > EXCEL_MAX_ROW:
        return False, f"{sheet}!{ref.ref} row exceeds Excel's maximum of {EXCEL_MAX_ROW}"
    if max_col is not None and max_col > EXCEL_MAX_COL:
        return False, f"{sheet}!{ref.ref} column exceeds Excel's maximum of {EXCEL_MAX_COL}"
    return True, None


def cells_from_locations(workbook, locations: Iterable[str]):
    for location in locations:
        if "!" not in location:
            continue
        sheet, coord = location.split("!", 1)
        if sheet in workbook.sheetnames:
            yield location, workbook[sheet][coord]
